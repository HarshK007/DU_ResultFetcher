from ResultFetcher import *
import openpyxl

def main():
    # print("Enter Your College/Department Name = ")
    clg_name = 'Department of Computer Science'
    print("Enter the Starting Enrollment Number of your Classmate = ")
    lwr = int(input())
    print("Enter the Ending Enrollment Number of your Classmate = ")
    upr = int(input())
    l1 = list(range(lwr, upr + 1))
    path = 'C:\\Users\\Harshit\\Documents\\MSCIresultFetcher.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "S.No."
    sheet.cell(row=1, column=2).value = "Enrollment No."
    sheet.cell(row=1, column=3).value = "Name"
    sheet.cell(row=1, column=4).value = "Total Credits"
    sheet.cell(row=1, column=5).value = "SGPA"
    sheet.cell(row=1, column=6).value = "Percentage"
    workbook.save(path)
    s_no = 1
    for x in l1:
        generate(s_no, clg_name, x)
        s_no = s_no + 1


if __name__ == '__main__':
    main()
