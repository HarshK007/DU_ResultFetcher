from selenium import webdriver
from WebTable import WebTable
from selenium.common.exceptions import NoSuchElementException
import openpyxl

def generate(sNo, clgName, rollNo):
    path = 'C:\\Users\\Harshit\\Documents\\MSCIresultFetcher.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=sNo+1, column=1).value = sNo
    driver = webdriver.Chrome("C:\\Users\\Harshit\\Downloads\\chromedriver_win32\\chromedriver.exe")
    url = "https://duresult.in/students/Home.aspx"
    driver.get(url)
    elem = driver.find_element_by_link_text("Statement of Marks").click()
    clgNameField = driver.find_element_by_id("ddlcollege").send_keys(clgName)
    rollNoField = driver.find_element_by_id("txtrollno").send_keys(rollNo)
    captchaImg = driver.find_element_by_id("imgCaptcha")
    captchaCode = captchaImg.get_attribute("src")
    captchaCodeStart = captchaCode.find("=") + 1
    captchaCodeEnd = captchaCode.find("&")
    captchaCode = captchaCode[captchaCodeStart:captchaCodeEnd]
    captchaField = driver.find_element_by_id("txtcaptcha").send_keys(captchaCode)
    printScoreCardBtn = driver.find_element_by_id("btnsearch").click()
    try:
        element2 = driver.find_element_by_id("lblrollno")
        sheet.cell(row=sNo+1, column=2).value = element2.text
        element2 = driver.find_element_by_id("lblname")
        sheet.cell(row=sNo + 1, column=3).value = element2.text
        #element = driver.find_element_by_id("gv_sgpa")
        #print(element.text)
        w = WebTable(driver.find_element_by_id('gv_sgpa'))
        row_count = w.get_row_count()
        total_sgpa = 0.0
        Total_credit = 0
        for i in range(1, row_count+1):
            row_data = w.row_data(i)
            #print(row_data)
            total_sgpa = total_sgpa + float(row_data[3])
            Total_credit = Total_credit + int(row_data[2])
        #print("TOTAL is = " + str(total_sum))
        sheet.cell(row=sNo + 1, column=4).value = Total_credit
        avgSGPA = total_sgpa/row_count
        sheet.cell(row=sNo+1, column=5).value = avgSGPA
        sheet.cell(row=sNo+1, column=6).value = (avgSGPA*9.5)
        #print("************************************************************************")
        workbook.save(path)
        driver.close()
    except NoSuchElementException as exception:
        sheet.cell(row=sNo + 1, column=2).value = rollNo
        sheet.cell(row=sNo + 1, column=3).value = 'NA'
        sheet.cell(row=sNo + 1, column=4).value = 0
        sheet.cell(row=sNo + 1, column=5).value = 0
        sheet.cell(row=sNo + 1, column=6).value = 0
        workbook.save(path)
        print("NO RECORD FOUND FOR -> " + str(rollNo))
        driver.close()