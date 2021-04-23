from ResultFetcher import *
import openpyxl

def main():
    #print("Enter Your College/Department Name = ")
    clg_name = 'Department of Computer Science'
    # print("Enter the Starting Enrollment Number of your Classmate = ")
    lwr = 1894091
    # print("Enter the Ending Enrollment Number of your Classmate = ")
    upr = 1894135
    # ch = 'y'
    # print("Add the Defaulters Enrollment numbers (Press 'N' or 'n' once finished)")
    l1 = list(range(lwr, upr + 1))
    #l1.remove(1724545)
    #l1.remove(1893913)
    path='E:\\mca ducs\\MCA Results\\MSC-V-resultFetcher.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "S.No."
    sheet.cell(row=1, column=2).value = "Enrollment No."
    sheet.cell(row=1, column=3).value = "Name"
    sheet.cell(row=1, column=4).value = "Total Marks"
    sheet.cell(row=1, column=5).value = "Percentage"
    sheet.cell(row=1, column=6).value = "Last Semester"
    sheet.cell(row=1, column=7).value = "Last Sem Percentage"
    workbook.save(path)
    s_no = 1
    for x in l1:
        generate(s_no, clg_name, x)
        s_no = s_no + 1


if __name__ == '__main__':
    main()
