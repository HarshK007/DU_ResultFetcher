from selenium import webdriver
from WebTable import WebTable
from selenium.common.exceptions import NoSuchElementException
import openpyxl

def generate(sNo, clgName, rollNo):
    path = 'E:\\mca ducs\\MCA Results\\MSC-V-resultFetcher.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=sNo+1, column=1).value = sNo
    driver = webdriver.Chrome("C:\\Users\\Dell\\Downloads\\chromedriver_win32\\chromedriver.exe")
    url = "http://durslt.du.ac.in/DURSLT_ND2020/Students/Home.aspx"
    driver.get(url)
    elem = driver.find_element_by_link_text("Statement of Marks").click()
    #clgName = 'Department of Computer Science'
    #rollNo = '1894105'
    clgNameField = driver.find_element_by_id("ddlcollege").send_keys(clgName)
    rollNoField = driver.find_element_by_id("txtrollno").send_keys(rollNo)
    captchaImg = driver.find_element_by_id("imgCaptcha")
    captchaCode = captchaImg.get_attribute("src")
    captchaCodeStart = captchaCode.find("=") + 1
    captchaCodeEnd = captchaCode.find("&")
    captchaCode = captchaCode[captchaCodeStart:captchaCodeEnd]
    captchaField = driver.find_element_by_id("txtcaptcha").send_keys(captchaCode)
    printScoreCardBtn = driver.find_element_by_id("btnsearch").click()
    try:
        element1 = driver.find_element_by_id("Label12")
        element2 = driver.find_element_by_id("lblrollno")
        #print(element1.text + "  " + element2.text)
        sheet.cell(row=sNo+1, column=2).value = element2.text
        element1 = driver.find_element_by_id("Label6")
        element2 = driver.find_element_by_id("lblname")
        #print(element1.text + "  " + element2.text)
        sheet.cell(row=sNo + 1, column=3).value = element2.text
        element = driver.find_element_by_id("gvrslt")
        #print(element.text)
        w = WebTable(driver.find_element_by_id('gvrslt'))
        row_count = w.get_row_count()
        total_sum = 0
        curr_Obtained = 0
        curr_total = 0
        weight = 0
        for i in range(1, row_count+1):
            row_data = w.row_data(i)
            curr_Obtained = int(row_data[1])
            curr_total = int(row_data[2])
            total_sum = total_sum + curr_Obtained
            weight = weight + curr_total
        #print("TOTAL is = " + str(total_sum))
        sheet.cell(row=sNo + 1, column=4).value = total_sum
        avg = total_sum/weight
        sheet.cell(row=sNo+1, column=5).value = (avg*100)
        sheet.cell(row=sNo + 1, column=6).value = curr_Obtained
        sheet.cell(row=sNo + 1, column=7).value = (curr_Obtained/curr_total)*100
        #print("************************************************************************")
        workbook.save(path)
        driver.close()
    except NoSuchElementException as exception:
        sheet.cell(row=sNo + 1, column=2).value = rollNo
        sheet.cell(row=sNo + 1, column=3).value = 'NA'
        sheet.cell(row=sNo + 1, column=4).value = 'NA'
        sheet.cell(row=sNo + 1, column=5).value = 'NA'
        workbook.save(path)
        print("NO RECORD FOUND FOR -> " + str(rollNo))
        driver.close()